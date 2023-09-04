import openpyxl
import json


def read_config(path):
    data_book = openpyxl.load_workbook('./' + path, read_only=True)
    value1 = read1(data_book.worksheets[0])
    value2 = read1(data_book.worksheets[1])
    value3 = read1(data_book.worksheets[2])

    result = {
        "prompts": value1,
        "fields": value2,
        "verification": value3,
        "page_config": [{
            "url_reg": "plcm.cncc.cn:30080/osc/.*",
            "page_type": "normal",
            "init_type": "observer",
            "form_container": {
                "locate_method": "querySelector",
                "location": "#view-item-detail-drawer"
            },
            "condition_for_pageload": {
                "locate_method": "querySelector",
                "location":"[data-element-id*='layout.content']"
            },
            "condition_for_observe_parentnode": {
                "locate_method": "querySelector",
                "location": ".ant-layout-content"
            },
            "condition_for_tooltip_label": {
                "locations": ["#static-scroll-container1"]
            }
        }, {
            "url_reg": "plcm.cncc.cn:30080/_team/.*",
            "page_type": "normal",
            "init_type": "observer",
            "form_container": {
                "locate_method": "querySelector",
                "location": "#static-scroll-container1"
            },
            "condition_for_pageload": {
                "locate_method": "querySelector",
                "location":"#static-scroll-container"
            },
            "condition_for_observe_parentnode": {
                "locate_method": "querySelector",
                "location": "#static-scroll-container"
            },
            "condition_for_tooltip_label": {
                "locations": ["#static-scroll-container1"]
            }
        }],

    }
    with open("src/assets/config/config.json", "w+", encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False)


def read1(sheet):
    ############## 提示信息 ################
    # 表头
    row = 2
    col = 1
    fieldsKey = []
    while True:
        value = sheet.cell(row, col).value
        if value in ["", None]:
            break
        fieldsKey.append(sheet.cell(row, col).value)
        col += 1
    # 内容
    fieldsContent = []
    row = 3
    while True:
        if sheet.cell(row, 1).value in ["", None]:
            break
        # 对于每一行数据，与表头组成key-value的形式，对于表头存在重复的，value变为数组
        obj = {}
        for col in range(1, len(fieldsKey)+1):
            key = fieldsKey[col-1]
            value = sheet.cell(row, col).value
            if key not in obj:
                obj[key] = value
            else:
                if type(obj[key]) != type([]):
                    obj[key] = [obj[key]]
                obj[key].append(value)
        row += 1
        print(obj)

        for k, v in obj.copy().items():
            # 对于复合元素（表头含有：的字段，values是一个数组），修改表示方式为 prefix:[{body1:'', body2: ''},{body1:'', body2: ''}] 的形式
            if ":" in k:
                prefix, body = k.split(":")
                # 初始化，key为prefix，value为n个空对象的数组
                if prefix not in obj:
                    obj[prefix] = [{} for _ in range(len(v))]
                # 将原来数组中的每个值，分别以key为body，value为该值的形式，放到对象中
                for idx, item in enumerate(v):
                    obj[prefix][idx][body] = item
                # 删除obj中的复合元素
                del (obj[k])
            elif k.startswith("L_") and v:
                # 对于数组元素（表头以L_开头的字段，以逗号分割字符串为数组）
                vlist = v.split(",")
                obj[k[2:]] = vlist
                # 删除obj中原来的元素
                del (obj[k])
            elif k.startswith("O_") and v:
                # 对于对象元素（表头以O_开头的字段，作为对象直接填入json）
                vdict = json.loads(v)
                obj[k[2:]] = vdict
                # 删除obj中原来的元素
                del (obj[k])

        # for k in keysToDel:
        fieldsContent.append(obj)

    return fieldsContent



if __name__ == '__main__':
    read_config('配置.xlsx')
